import time
import calendar
import pyautogui
import getpass
import openpyxl as op
import math

import selenium.common.exceptions
from openpyxl import Workbook
from datetime import datetime, timedelta

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime

class Workday:
    def __init__(self):
        user = input()
        password = getpass.getpass("Password:")
        self.driver = webdriver.Chrome()
        self.loginToWorkday(user, password)
        pass

    def loginToWorkday(self,user,password):
        print("Logging into workday as '%s'"%(user))
        d = self.driver
        d.get("https://wd3.myworkday.com/polestar/d/home.htmld") # Production
        # d.get("https://wd3-impl.workday.com/polestar2/d/home.htmld") # Testing T2
        WebDriverWait(d,30).until(EC.presence_of_element_located((By.XPATH,'//div[contains(text(), "SSO")]')))
        d.find_element(By.XPATH,'//div[contains(text(), "SSO")]').click()

        # ID
        WebDriverWait(d,30).until(EC.presence_of_element_located((By.NAME, "loginfmt")))
        d.find_element(By.NAME, "loginfmt").send_keys(user)
        time.sleep(0.5)
        WebDriverWait(d, 30).until(EC.presence_of_element_located((By.XPATH, '//input[@value = "Next"]')))
        d.find_element(By.XPATH, '//input[@value = "Next"]').click()

        # Password
        time.sleep(1)
        WebDriverWait(d,30).until(EC.presence_of_element_located((By.NAME, "passwd")))
        d.find_element(By.NAME, "passwd").send_keys(password)
        time.sleep(0.5)
        WebDriverWait(d,30).until(EC.presence_of_element_located((By.XPATH, '//input[@value = "Sign in"]')))
        d.find_element(By.XPATH, '//input[@value = "Sign in"]').click()

        # Authentication
        time.sleep(1)
        WebDriverWait(d,30).until(EC.presence_of_element_located((By.CLASS_NAME, "displaySign")))
        print(d.find_element(By.CLASS_NAME, "displaySign").text)

        # Finish
        time.sleep(1)
        WebDriverWait(d,30).until(EC.presence_of_element_located((By.XPATH, '//*[contains(text(), "Stay signed in?")]')))
        d.find_element(By.XPATH, '//input[@value = "Yes"]').click()

        time.sleep(3)
        WebDriverWait(d,30).until(EC.presence_of_element_located((By.ID, 'tdCheckbox')))
        d.find_element(By.ID, 'tdCheckbox').click()
        d.find_element(By.ID, 'submitButton').click()

        time.sleep(3)
        WebDriverWait(d,30).until(EC.presence_of_element_located((By.CLASS_NAME, "wdappchrome-ab")))
        time.sleep(3)
        print("Successfully login Workday")

        return d

    def new_tab(self):
        d = self.driver
        d.switch_to.new_window()
        d.get("https://wd3.myworkday.com/polestar/d/home.htmld")
        WebDriverWait(d,30).until(EC.presence_of_element_located((By.CLASS_NAME, "wdappchrome-ab")))
        time.sleep(3)
        return d

    def wdreport_findsupplierinvoices(self):
        d = self.driver
        action = ActionChains(d)
        d.find_element(By.XPATH, '//input[@data-automation-id ="globalSearchInput"]').click()
        time.sleep(1)
        action.send_keys("find supplier invoices").perform()
        time.sleep(3)
        action.send_keys(Keys.ARROW_DOWN).send_keys(Keys.ENTER).perform()
        time.sleep(3)

        ##### Input Parameter #####
        d.find_element(By.XPATH, '//label[contains(text(), "Company")]').click()
        time.sleep(1)
        action.send_keys("KR21").send_keys(Keys.ENTER).send_keys(Keys.TAB).perform()
        time.sleep(3)

        d.find_element(By.XPATH, '//label[contains(text(), "Invoice Status")]').click()
        time.sleep(1)
        action.send_keys("Approved").send_keys(Keys.ENTER).perform()
        time.sleep(1.5)
        action.send_keys("Draft").send_keys(Keys.ENTER).perform()
        time.sleep(1.5)
        action.send_keys("In Progress").send_keys(Keys.ENTER).send_keys(Keys.TAB).perform()
        time.sleep(1.5)

        d.find_element(By.XPATH, '//label[contains(text(), "Invoice Date On or After")]').click()
        time.sleep(1)
        action.send_keys("0", "1", "0", "1", "2", "0", "2", "3").send_keys(Keys.TAB).perform()
        time.sleep(3)

        d.find_element(By.XPATH, '//label[contains(text(), "Payment Status")]').click()
        time.sleep(1)
        action.send_keys("Partially Paid").send_keys(Keys.ENTER).perform()
        time.sleep(1.5)
        action.send_keys("Unpaid").send_keys(Keys.ENTER).send_keys(Keys.TAB).perform()
        time.sleep(1.5)

        d.find_element(By.XPATH, '//span[@title = "OK"]').click()
        WebDriverWait(d,30).until(EC.presence_of_element_located((By.XPATH, './/div[@id="4311$94465"]')))

        excel_row_num = 2
        cnt_items = d.find_element(By.XPATH, './/label[@data-automation-id="rowCountLabel"]').text
        cnt_items = int(cnt_items.strip(' /items')) - 1
        total_page_num = math.ceil(cnt_items / 30)

        # Add a script for copying row[0] as a report header line, this is one time script not for loop

        for page_num in range(1, total_page_num + 1):

            if page_num == 1:
                pass
            else:
                d.find_element(By.XPATH, f'.//button[@data-automation-id="page{page_num}"]').click()

            rows = d.find_elements(By.TAG_NAME, 'tr')

            for row in rows[1:]:
                # Variable values
                workbook.active.cell(row=excel_row_num, column=1).value = row.find_element(By.XPATH,'.//div[@id="4311$94465"]').text
                workbook.active.cell(row=excel_row_num, column=3).value = row.find_element(By.XPATH,'.//div[@id="4311$94464"]').text
                workbook.active.cell(row=excel_row_num, column=5).value = row.find_element(By.XPATH,'.//div[@id="4311$94470"]').text
                workbook.active.cell(row=excel_row_num, column=6).value = row.find_element(By.XPATH,'.//div[@id="4311$94460"]').text
                workbook.active.cell(row=excel_row_num, column=7).value = row.find_element(By.XPATH,'.//div[@id="4311$94455"]').text
                workbook.active.cell(row=excel_row_num, column=9).value = row.find_element(By.XPATH,'.//div[@id="4311$94459"]').text
                workbook.active.cell(row=excel_row_num, column=10).value = row.find_element(By.XPATH,'.//div[@id="4311$94452"]').text
                workbook.active.cell(row=excel_row_num, column=11).value = row.find_element(By.XPATH,'.//div[@id="4311$94451"]').text
                workbook.active.cell(row=excel_row_num, column=12).value = row.find_element(By.XPATH,'.//div[@id="4311$94469"]').text
                workbook.active.cell(row=excel_row_num, column=15).value = row.find_element(By.XPATH,'.//div[@id="4311$94457"]').text
                workbook.active.cell(row=excel_row_num, column=16).value = row.find_element(By.XPATH,'.//div[@id="4311$94461"]').text

                # Fixed values
                workbook.active.cell(row=excel_row_num, column=2).value = "KR21 Polestar Automotive Korea Limited"

                # Determine Supplier category in Column Q
                if "SE02" in workbook.active.cell(row=excel_row_num, column=5).value:
                    workbook.active.cell(row=excel_row_num, column=17).value = "PPAB"
                elif "Volvo Car" in workbook.active.cell(row=excel_row_num, column=5).value:
                    workbook.active.cell(row=excel_row_num, column=17).value = "Related Party"
                else:
                    workbook.active.cell(row=excel_row_num, column=17).value = "3rd Party"

                # Determine future payment date in Column
                due_date = workbook.active.cell(row=excel_row_num, column=9).value
                due_date_object = datetime.strptime(due_date, "%d/%m/%Y")
                day_of_week = due_date_object.strftime("%A")
                payment_date = workbook.active.cell(row=excel_row_num, column=18)
                if day_of_week == "Monday" or "Wednesday":
                    payment_date.value = due_date_object.strftime('%d/%m/%Y')
                if day_of_week == "Tuesday" or "Thursday":
                    payment_date.value = (due_date_object - timedelta(days=1)).strftime('%d/%m/%Y')
                if day_of_week == "Friday":
                    payment_date.value = (due_date_object - timedelta(days=2)).strftime('%d/%m/%Y')
                if day_of_week == "Saturday":
                    payment_date.value = (due_date_object - timedelta(days=3)).strftime('%d/%m/%Y')
                if day_of_week == "Sunday":
                    payment_date.value = (due_date_object - timedelta(days=4)).strftime('%d/%m/%Y')

                excel_row_num = excel_row_num + 1

            time.sleep(5)

    def wdreport_payableaging(self):
        d = self.driver
        action = ActionChains(d)
        d.find_element(By.XPATH, '//input[@data-automation-id ="globalSearchInput"]').click()
        time.sleep(1)
        action.send_keys("Payables Aging").perform()
        time.sleep(3)
        action.send_keys(Keys.ARROW_DOWN).send_keys(Keys.ENTER).perform()
        time.sleep(3)

        ##### Input Parameter #####
        d.find_element(By.XPATH, '//label[contains(text(), "Company")]').click()
        time.sleep(1)
        action.send_keys("KR21").send_keys(Keys.ENTER).send_keys(Keys.TAB).perform()
        time.sleep(3)

        d.find_element(By.XPATH, '//label[contains(text(), "Aging Group")]').click()
        time.sleep(1)
        action.send_keys("standard").send_keys(Keys.ENTER).send_keys(Keys.TAB).perform()
        time.sleep(3)

        d.find_element(By.XPATH, '//span[@title = "OK"]').click()
        WebDriverWait(d,30).until(EC.presence_of_element_located((By.XPATH, './/div[@id="56$675250"]')))

        # Add FOR loop script to handle multiple pages when needed

        rows = d.find_elements(By.TAG_NAME, 'tr')

        excel_row_num = 1
        for row in rows:
            line_item = row.text
            line_item = line_item.split('\n')
            for column in range(0, len(line_item)):
                workbook.active.cell(row=excel_row_num, column=column + 1).value = line_item[column]
            excel_row_num += 1





wd = Workday()

workbook = Workbook()
active_sheet = workbook.active
active_sheet.title = "find supplier invoices"
print("preparing invoice report...")

try:
    wd.wdreport_findsupplierinvoices()
except:
    pass

print("invoice report copied!")
print("preparing aging report...")
time.sleep(3)
wd.new_tab()

new_sheet = workbook.create_sheet("payable aging")
workbook.active = new_sheet
wd.wdreport_payableaging()
print("aging report copied!")

current_time = datetime.now()
file_location = "C:/Users/Public/"
file_name = f"{file_location}kr21_ptp_aging_{current_time.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
workbook.save(file_name)
workbook.close()
print("ptp report completed")
